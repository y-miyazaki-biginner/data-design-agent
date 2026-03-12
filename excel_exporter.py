"""Excel手順書エクスポーター - openpyxlで3シート構成のExcelを生成"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# スタイル定義
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Yu Gothic", size=10)
BOLD_FONT = Font(name="Yu Gothic", size=10, bold=True)
TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True, color="2D3748")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
STEP_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
VALIDATION_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
FLOW_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")


def _apply_header_style(ws, row, col_count):
    """ヘッダー行にスタイルを適用"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _apply_cell_style(cell, wrap=False, bold=False):
    """セルにスタイルを適用"""
    cell.font = BOLD_FONT if bold else CELL_FONT
    cell.alignment = WRAP_ALIGN if wrap else Alignment(vertical="top")
    cell.border = THIN_BORDER


def export_excel(result: dict) -> bytes:
    """生成結果をExcel形式でエクスポートする"""
    wb = Workbook()

    # === シート1: データ設計 ===
    ws1 = wb.active
    ws1.title = "データ設計"
    _write_design_sheet(ws1, result.get("data_design", []))

    # === シート2: 実装手順 ===
    ws2 = wb.create_sheet("実装手順")
    _write_steps_sheet(ws2, result.get("implementation_steps", []))

    # === シート3: データフロー ===
    ws3 = wb.create_sheet("データフロー")
    _write_flow_sheet(ws3, result.get("data_flow", []))

    # バイトに変換
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _write_design_sheet(ws, data_design: list):
    """シート1: データ設計テーブル"""
    # タイトル
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "データ設計書"
    title_cell.font = TITLE_FONT

    # ヘッダー
    headers = ["ソーステーブル", "カラム論理名", "カラム物理名", "推定型", "NULL許可", "備考"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))

    # データ行
    for item in data_design:
        row += 1
        values = [
            item.get("source_table", ""),
            item.get("logical_name", ""),
            item.get("physical_name", ""),
            item.get("estimated_type", ""),
            "NULLABLE" if item.get("nullable") else "NOT NULL",
            item.get("note", ""),
        ]
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=i, value=v)
            _apply_cell_style(cell, wrap=(i == 6))

    # カラム幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 30


def _write_steps_sheet(ws, steps: list):
    """シート2: 実装手順"""
    # タイトル
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "b→dash 実装手順書"
    title_cell.font = TITLE_FONT

    row = 3

    for step in steps:
        step_num = step.get("step_number", "")
        title = step.get("title", "")
        desc = step.get("description", "")
        bdash = step.get("bdash_operation", {})
        val = step.get("validation", {})

        # ステップヘッダー
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=f"ステップ {step_num}: {title}")
        cell.font = Font(name="Yu Gothic", size=12, bold=True, color="1A237E")
        cell.fill = STEP_FILL
        cell.border = THIN_BORDER
        for c in range(2, 5):
            ws.cell(row=row, column=c).fill = STEP_FILL
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

        # 説明
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=desc)
        _apply_cell_style(cell, wrap=True)
        row += 1

        # GUI操作手順
        gui_steps = bdash.get("gui_steps", [])
        if gui_steps:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value="【b→dash GUI操作手順】")
            cell.font = Font(name="Yu Gothic", size=10, bold=True, color="4472C4")
            row += 1

            for i, gs in enumerate(gui_steps, 1):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                cell = ws.cell(row=row, column=1, value=f"  {i}. {gs}")
                _apply_cell_style(cell, wrap=True)
                row += 1

        # 検証観点
        if val:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            cell = ws.cell(row=row, column=1, value="【検証観点】")
            cell.font = Font(name="Yu Gothic", size=10, bold=True, color="E65100")
            cell.fill = VALIDATION_FILL
            for c in range(2, 5):
                ws.cell(row=row, column=c).fill = VALIDATION_FILL
            row += 1

            val_items = [
                ("確認内容", val.get("what_to_check", "")),
                ("確認手順", val.get("how_to_check", "")),
                ("期待値", val.get("expected_result", "")),
                ("NG時対処", val.get("ng_action", "")),
                ("顧客説明", val.get("customer_explanation", "")),
            ]
            for label, value in val_items:
                if value:
                    cell_label = ws.cell(row=row, column=1, value=label)
                    _apply_cell_style(cell_label, bold=True)
                    cell_label.fill = VALIDATION_FILL
                    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                    cell_val = ws.cell(row=row, column=2, value=value)
                    _apply_cell_style(cell_val, wrap=True)
                    cell_val.fill = VALIDATION_FILL
                    row += 1

        # 空行
        row += 1

    # カラム幅
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30


def _write_flow_sheet(ws, data_flow: list):
    """シート3: データフロー図"""
    # タイトル
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = "データフロー図"
    title_cell.font = TITLE_FONT

    # ヘッダー
    headers = ["ステップ", "入力データ", "操作内容", "出力データ"]
    row = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
    _apply_header_style(ws, row, len(headers))

    # フローデータ
    for i, flow in enumerate(data_flow, 1):
        row += 1
        from_data = " + ".join(flow.get("from", []))
        operation = flow.get("operation", "")
        to_data = flow.get("to", "")

        cell_step = ws.cell(row=row, column=1, value=i)
        _apply_cell_style(cell_step)
        cell_step.alignment = CENTER_ALIGN

        cell_from = ws.cell(row=row, column=2, value=from_data)
        _apply_cell_style(cell_from, wrap=True)
        cell_from.fill = FLOW_FILL

        cell_op = ws.cell(row=row, column=3, value=operation)
        _apply_cell_style(cell_op, wrap=True)

        cell_to = ws.cell(row=row, column=4, value=to_data)
        _apply_cell_style(cell_to, wrap=True)
        cell_to.fill = FLOW_FILL

    # フロー図（テキストベース）
    if data_flow:
        row += 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1, value="【フロー概要】").font = Font(name="Yu Gothic", size=11, bold=True)
        row += 1

        flow_text = _build_flow_text(data_flow)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 5, end_column=4)
        cell = ws.cell(row=row, column=1, value=flow_text)
        cell.font = Font(name="Consolas", size=10)
        cell.alignment = WRAP_ALIGN

    # カラム幅
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 25


def _build_flow_text(data_flow: list) -> str:
    """データフローのテキスト表現を生成"""
    lines = []
    for i, flow in enumerate(data_flow):
        from_data = " + ".join(flow.get("from", []))
        operation = flow.get("operation", "")
        to_data = flow.get("to", "")

        if i > 0:
            lines.append("        |")
            lines.append("        v")
        lines.append(f"  {from_data}")
        lines.append(f"    --[{operation}]-->")
        lines.append(f"  {to_data}")

    return "\n".join(lines)
